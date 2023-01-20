from bot.constants import DRIVER_URL, WEB_URL, FILE_URL
from openpyxl import load_workbook
from datetime import date
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.options import Options
import time
import os
class GoogleBot(webdriver.Chrome):

    def __init__(self, driver_path=DRIVER_URL, teardown=False):
        self.driver_path = driver_path
        self.teardown = teardown
        os.environ['PATH'] += self.driver_path
        super(GoogleBot, self).__init__()
        self.implicitly_wait(15)  # wait for 15 seconds for current and next find methods
        self.maximize_window()
        self.keywords = []
        self.wb = wb = load_workbook(FILE_URL)
        self.ws = None

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.teardown:
            self.quit()

    def checkDay(self):
        if date.today().weekday() == 0:
            return "Monday"
        if date.today().weekday() == 1:
            return "Tuesday"
        if date.today().weekday() == 2:
            return "Wednesday"
        if date.today().weekday() == 3:
            return "Thursday"
        if date.today().weekday() == 4:
            return "Friday"
        if date.today().weekday() == 5:
            return "Saturday"
        if date.today().weekday() == 6:
            return "Sunday"

    def loadKeywords(self):
        day = self.checkDay()
        self.ws = self.wb[day]
        index = 1

        for col in self.ws['C']:
            if index >= 3 and index <= 12:
                self.keywords.append(col.value)
            index += 1

    def loadExcelFile(self):
        wb = load_workbook('Excel.xlsx')

    def run(self):
        self.loadExcelFile()
        self.loadKeywords()

        self.get(WEB_URL)

        search_element = self.find_element(
            By.NAME,
            'q'
        )
        cell_index = 3

        print(self.keywords)

        # go through all the keywords in the excel file for particular day
        for i in range(len(self.keywords)):
            time.sleep(3)
            search_element.clear()
            search_element.send_keys(self.keywords[i])
            time.sleep(3)

            # get all the elements from listbox of auto suggestion
            sug_list = self.find_elements(By.XPATH, "//*[@role='listbox']/li")


            list = []
            # append it to a list
            for item in sug_list:
                list.append(item)

            text_list = []
            # convert it to text list
            for item in list:
                x = item.text
                if '\n' in x:
                    # if new line is there take the upper element rather then taking the lower one
                    text_list.append(x.split('\n')[0])
                else:
                    text_list.append(x)
            print(text_list)

            # find out the shortest and longest text
            shortest = min(text_list, key=len)
            longest = max(text_list, key=len)

            # write it to excel file
            self.ws[f'D{str(cell_index)}'] = longest
            self.ws[f'E{str(cell_index)}'] = shortest
            self.wb.save("Excel.xlsx")

            print(shortest);
            print(longest);

            # check if it is in the end of excel file
            if cell_index < 12:
                cell_index += 1
            #print(self.keywords[i])









